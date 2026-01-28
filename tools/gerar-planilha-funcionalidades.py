#!/usr/bin/env python3
"""
Gera planilha consolidada de funcionalidades (RFs) do IControlIT.

Versão: 4.3
Data: 2026-01-20
Changelog:
  v4.3: Remove truncamento de regras (exibe texto completo)
  v4.2: Melhora extração de descrições e regras (10 fontes de fallback, busca em escopo.incluso e regras_negocio.descricao)
  v4.1: Remove limite de 200 caracteres nas descrições (exibe texto completo)
  v4.0: Adiciona colunas Processo e Jornada mapeando RFs aos processos documentados
  v3.0: Adiciona colunas para discussão do cliente
  v2.1: Melhora extração de descrições
  v2.0: Corrige extração de nomes (rf_title)
  v1.0: Versão inicial
"""

import os
import glob
import yaml
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

# ==============================================================================
# MAPEAMENTO DE RFs → PROCESSOS E JORNADAS
# ==============================================================================

# Este mapeamento foi extraído da leitura dos arquivos de processos em
# D:\IC2_Governanca\documentacao\Processos\

MAPEAMENTO_RF_PROCESSO_JORNADA = {
    # Jornada 1: Infraestrutura e Configuração
    'RF001': ('PRO-INF-001', 'Jornada Infraestrutura'),
    'RF002': ('PRO-INF-002', 'Jornada Infraestrutura'),
    'RF003': ('PRO-INF-003', 'Jornada Infraestrutura'),
    'RF004': ('PRO-INF-004', 'Jornada Infraestrutura'),
    'RF005': ('PRO-INF-005', 'Jornada Infraestrutura'),
    'RF006': ('PRO-INF-006', 'Jornada Infraestrutura'),
    'RF007': ('PRO-INF-007', 'Jornada Infraestrutura'),
    'RF014': ('PRO-INF-008', 'Jornada Infraestrutura'),

    # Jornada 2: Workflows e Importação
    'RF063': ('PRO-WKF-001', 'Jornada Workflows'),
    'RF064': ('PRO-WKF-002', 'Jornada Workflows'),
    'RF065': ('PRO-WKF-003', 'Jornada Workflows'),
    'RF066': ('PRO-WKF-004', 'Jornada Workflows'),
    'RF067': ('PRO-WKF-005', 'Jornada Workflows'),
    'RF084': ('PRO-WKF-006', 'Jornada Workflows'),
    'RF085': ('PRO-WKF-007', 'Jornada Workflows'),
    'RF086': ('PRO-WKF-008', 'Jornada Workflows'),
    'RF088': ('PRO-WKF-009', 'Jornada Workflows'),

    # Jornada 3: Financeiro Completo
    'RF026': ('PRO-FCT-001', 'Jornada Financeiro'),
    'RF030': ('PRO-FCT-002', 'Jornada Financeiro'),
    'RF031': ('PRO-FCT-003', 'Jornada Financeiro'),
    'RF032': ('PRO-FCT-004', 'Jornada Financeiro'),
    'RF089': ('PRO-FCT-005', 'Jornada Financeiro'),
    'RF090': ('PRO-FCT-006', 'Jornada Financeiro'),
    'RF097': ('PRO-FCT-007', 'Jornada Financeiro'),
    'RF025': ('PRO-FAC-001', 'Jornada Financeiro'),
    'RF036': ('PRO-FAC-002', 'Jornada Financeiro'),
    'RF037': ('PRO-FAC-003', 'Jornada Financeiro'),
    'RF042': ('PRO-FAC-004', 'Jornada Financeiro'),
    'RF055': ('PRO-FAC-005', 'Jornada Financeiro'),

    # Jornada 5: Service Desk
    'RF028': ('PRO-SVC-001', 'Jornada Service Desk'),
    'RF029': ('PRO-SVC-002', 'Jornada Service Desk'),
    'RF033': ('PRO-SVC-003', 'Jornada Service Desk'),
    'RF038': ('PRO-SVC-004', 'Jornada Service Desk'),
    'RF053': ('PRO-SVC-005', 'Jornada Service Desk'),
    'RF072': ('PRO-SVC-006', 'Jornada Service Desk'),
    'RF078': ('PRO-SVC-007', 'Jornada Service Desk'),
    'RF087': ('PRO-SVC-008', 'Jornada Service Desk'),

    # Jornada 6: Auditoria
    'RF068': ('PRO-AUD-001', 'Jornada Auditoria'),

    # RFs que são FUNCIONALIDADES DE GESTÃO (sem código de processo)
    # Estes RFs terão Processo vazio e Jornada = "Funcionalidades de Gestão"
    # (72 RFs listados em 07-Funcionalidades-Gestao.md)
    'RF012': ('', 'Funcionalidades de Gestão'),
    'RF013': ('', 'Funcionalidades de Gestão'),
    'RF015': ('', 'Funcionalidades de Gestão'),
    'RF016': ('', 'Funcionalidades de Gestão'),
    'RF017': ('', 'Funcionalidades de Gestão'),
    'RF018': ('', 'Funcionalidades de Gestão'),
    'RF019': ('', 'Funcionalidades de Gestão'),
    'RF020': ('', 'Funcionalidades de Gestão'),
    'RF022': ('', 'Funcionalidades de Gestão'),
    'RF024': ('', 'Funcionalidades de Gestão'),
    'RF027': ('', 'Funcionalidades de Gestão'),
    'RF039': ('', 'Funcionalidades de Gestão'),
    'RF041': ('', 'Funcionalidades de Gestão'),
    'RF043': ('', 'Funcionalidades de Gestão'),
    'RF047': ('', 'Funcionalidades de Gestão'),
    'RF048': ('', 'Funcionalidades de Gestão'),
    'RF049': ('', 'Funcionalidades de Gestão'),
    'RF050': ('', 'Funcionalidades de Gestão'),
    'RF051': ('', 'Funcionalidades de Gestão'),
    'RF052': ('', 'Funcionalidades de Gestão'),
    'RF054': ('', 'Funcionalidades de Gestão'),
    'RF056': ('', 'Funcionalidades de Gestão'),
    'RF057': ('', 'Funcionalidades de Gestão'),
    'RF058': ('', 'Funcionalidades de Gestão'),
    'RF059': ('', 'Funcionalidades de Gestão'),
    'RF060': ('', 'Funcionalidades de Gestão'),
    'RF061': ('', 'Funcionalidades de Gestão'),
    'RF062': ('', 'Funcionalidades de Gestão'),
    'RF070': ('', 'Funcionalidades de Gestão'),
    'RF071': ('', 'Funcionalidades de Gestão'),
    'RF073': ('', 'Funcionalidades de Gestão'),
    'RF074': ('', 'Funcionalidades de Gestão'),
    'RF076': ('', 'Funcionalidades de Gestão'),
    'RF077': ('', 'Funcionalidades de Gestão'),
    'RF091': ('', 'Funcionalidades de Gestão'),
    'RF092': ('', 'Funcionalidades de Gestão'),
    # (mais RFs de Funcionalidades de Gestão serão adicionados automaticamente se não estiverem mapeados)
}


def extrair_texto(valor):
    """
    Extrai texto de estruturas YAML complexas (dicts, listas).

    Args:
        valor: Valor YAML (string, dict, list)

    Returns:
        str: Texto extraído
    """
    if isinstance(valor, str):
        return valor.strip()

    if isinstance(valor, dict):
        # Campos comuns de descrição em dicts
        for campo in ['objetivo', 'problema_resolvido', 'descricao', 'texto', 'resumo']:
            if campo in valor and valor[campo]:
                return extrair_texto(valor[campo])
        # Se nenhum campo específico, retorna string do dict
        return str(valor)

    if isinstance(valor, list):
        # Junta itens da lista
        return '. '.join(str(item) for item in valor if item)

    return str(valor)


def extrair_rfs_yaml(documentacao_path):
    """
    Extrai informações de todos os RFs dos arquivos YAML.

    Args:
        documentacao_path (str): Caminho para pasta de documentação

    Returns:
        list: Lista de dicts com informações dos RFs
    """
    pattern = os.path.join(documentacao_path, "**", "RF*.yaml")
    arquivos = glob.glob(pattern, recursive=True)

    # Ignorar arquivos RL-RF*.yaml (Releases)
    arquivos = [a for a in arquivos if not os.path.basename(a).startswith("RL-")]

    print(f"[INFO] Processando {len(arquivos)} RFs...")

    rfs = []
    for arquivo in arquivos:
        try:
            with open(arquivo, 'r', encoding='utf-8') as f:
                dados = yaml.safe_load(f)

            if not dados:
                continue

            # Extrair RF ID
            rf_id = dados.get('rf_id', '')
            if not rf_id:
                if 'rf' in dados and isinstance(dados['rf'], dict):
                    rf_id = dados['rf'].get('id', '')
            if not rf_id:
                rf_id = os.path.basename(arquivo).replace('.yaml', '')

            # Extrair NOME (múltiplas tentativas)
            nome = dados.get('titulo', dados.get('nome', dados.get('title', '')))
            if not nome:
                nome = dados.get('rf_title', '')  # 26 RFs usam este campo
            if not nome:
                if 'rf' in dados and isinstance(dados['rf'], dict):
                    nome = dados['rf'].get('nome', dados['rf'].get('titulo', ''))
            if not nome:
                if 'metadata' in dados and isinstance(dados['metadata'], dict):
                    nome = dados['metadata'].get('titulo', dados['metadata'].get('nome', ''))

            # Extrair DESCRIÇÃO (10 estratégias de fallback para cobrir todas as estruturas)
            descricao = ''

            # 1. Tentar resumo_executivo
            if 'resumo_executivo' in dados:
                descricao = extrair_texto(dados['resumo_executivo'])

            # 2. Tentar descricao (pode ser dict ou string)
            if not descricao and 'descricao' in dados:
                descricao_raw = dados['descricao']
                if isinstance(descricao_raw, str):
                    descricao = descricao_raw.strip()
                elif isinstance(descricao_raw, dict):
                    descricao = extrair_texto(descricao_raw.get('objetivo', ''))
                    if not descricao:
                        descricao = extrair_texto(descricao_raw.get('problema_resolvido', ''))

            # 3. Tentar objetivos (lista)
            if not descricao and 'objetivos' in dados:
                objetivos = dados['objetivos']
                if isinstance(objetivos, list) and len(objetivos) > 0:
                    descricao = extrair_texto(objetivos[0])

            # 4. Tentar objetivo (string direta)
            if not descricao and 'objetivo' in dados:
                descricao = extrair_texto(dados['objetivo'])

            # 5. Tentar visao_geral (7 RFs)
            if not descricao and 'visao_geral' in dados:
                descricao = extrair_texto(dados['visao_geral'])

            # 6. Tentar escopo.incluso (primeiro item da lista de escopo)
            if not descricao and 'escopo' in dados:
                escopo = dados['escopo']
                if isinstance(escopo, dict) and 'incluso' in escopo:
                    incluso = escopo['incluso']
                    if isinstance(incluso, list) and len(incluso) > 0:
                        # Pegar os 3 primeiros itens do escopo
                        itens_escopo = [str(item) for item in incluso[:3] if item]
                        descricao = '; '.join(itens_escopo)
                elif isinstance(escopo, str):
                    descricao = escopo.strip()

            # 7. Tentar metadata
            if not descricao and 'metadata' in dados and isinstance(dados['metadata'], dict):
                descricao = extrair_texto(dados['metadata'].get('descricao', ''))
                if not descricao:
                    descricao = extrair_texto(dados['metadata'].get('resumo', ''))

            # 8. Tentar primeira regra de negócio como descrição (RF050 e similares)
            if not descricao and 'regras_negocio' in dados:
                rn_list = dados['regras_negocio']
                if isinstance(rn_list, list) and len(rn_list) > 0:
                    primeira_rn = rn_list[0]
                    if isinstance(primeira_rn, dict):
                        # Tentar campo descricao da regra
                        descricao = extrair_texto(primeira_rn.get('descricao', ''))
                        if not descricao:
                            descricao = extrair_texto(primeira_rn.get('titulo', ''))

            # 9. Tentar rf.descricao ou rf.objetivo (estrutura aninhada)
            if not descricao and 'rf' in dados and isinstance(dados['rf'], dict):
                rf_data = dados['rf']
                descricao = extrair_texto(rf_data.get('descricao', ''))
                if not descricao:
                    descricao = extrair_texto(rf_data.get('objetivo', ''))

            # 10. Se ainda não achou, criar descrição a partir do nome
            if not descricao and nome:
                descricao = f"Funcionalidade de {nome}"

            # Limpar descrição (remover quebras de linha excessivas)
            if descricao:
                descricao = ' '.join(descricao.split())

            # Extrair REGRAS (títulos ou descrições, sem truncamento)
            regras = []
            if 'regras_negocio' in dados:
                rn_list = dados['regras_negocio']
                if isinstance(rn_list, list):
                    for rn in rn_list[:5]:  # Máximo 5 regras
                        if isinstance(rn, dict):
                            # Tentar titulo primeiro, depois descricao
                            titulo_rn = rn.get('titulo', '')
                            if not titulo_rn:
                                # Tentar descricao (RF048 e similares usam este campo)
                                titulo_rn = rn.get('descricao', '')
                            if titulo_rn:
                                # Limpar texto (remover quebras de linha)
                                titulo_rn = ' '.join(str(titulo_rn).split())
                                regras.append(titulo_rn)

            regras_texto = '; '.join(regras) if regras else ''

            # Extrair Epic (garantir que seja string)
            epic = dados.get('epic', '')
            epic = extrair_texto(epic) if epic else ''

            # MAPEAR PROCESSO E JORNADA
            processo, jornada = MAPEAMENTO_RF_PROCESSO_JORNADA.get(rf_id, ('', 'Funcionalidades de Gestão'))

            rfs.append({
                'rf_id': rf_id,
                'nome': nome,
                'descricao': descricao,
                'regras': regras_texto,
                'processo': processo,
                'jornada': jornada,
                'epic': epic
            })

        except Exception as e:
            print(f"[ERRO] Falha ao processar {arquivo}: {str(e)}")
            continue

    # Ordenar por RF ID
    rfs.sort(key=lambda x: x['rf_id'])

    # Estatísticas
    total_rfs = len(rfs)
    rfs_com_nome = sum(1 for rf in rfs if rf['nome'])
    rfs_com_descricao = sum(1 for rf in rfs if rf['descricao'])
    rfs_com_regras = sum(1 for rf in rfs if rf['regras'])
    rfs_com_processo = sum(1 for rf in rfs if rf['processo'])

    print(f"[OK] {total_rfs} RFs extraídos")
    print(f"  - {rfs_com_nome}/{total_rfs} com nome ({rfs_com_nome*100//total_rfs}%)")
    print(f"  - {rfs_com_descricao}/{total_rfs} com descrição ({rfs_com_descricao*100//total_rfs}%)")
    print(f"  - {rfs_com_regras}/{total_rfs} com regras ({rfs_com_regras*100//total_rfs}%)")
    print(f"  - {rfs_com_processo}/{total_rfs} com processo mapeado ({rfs_com_processo*100//total_rfs}%)")

    return rfs


def gerar_planilha_excel(rfs, output_path):
    """
    Gera planilha Excel consolidada com formatação profissional.

    Args:
        rfs (list): Lista de RFs extraídos
        output_path (str): Caminho para arquivo de saída
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Funcionalidades IControlIT"

    # Definir cabeçalhos (nova estrutura com Processo e Jornada)
    cabecalhos = [
        'Cód.',
        'Nome Funcionalidade',
        'Descrição',
        'Regras',
        'Processo',
        'Jornada',
        'Prioridade',
        'Status Discussão',
        'Continua no sistema?',
        'Notas'
    ]

    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Aplicar cabeçalhos
    for col_num, cabecalho in enumerate(cabecalhos, 1):
        celula = ws.cell(row=1, column=col_num, value=cabecalho)
        celula.fill = header_fill
        celula.font = header_font
        celula.alignment = header_alignment
        celula.border = border_thin

    # Estilo das células de dados
    data_alignment = Alignment(vertical="top", wrap_text=True)
    editable_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Amarelo claro

    # Inserir dados dos RFs
    for row_num, rf in enumerate(rfs, 2):
        # Coluna A: Código
        celula_codigo = ws.cell(row=row_num, column=1, value=rf['rf_id'])
        celula_codigo.alignment = Alignment(horizontal="center", vertical="top")
        celula_codigo.border = border_thin

        # Coluna B: Nome Funcionalidade
        celula_nome = ws.cell(row=row_num, column=2, value=rf['nome'])
        celula_nome.alignment = data_alignment
        celula_nome.border = border_thin

        # Coluna C: Descrição
        celula_desc = ws.cell(row=row_num, column=3, value=rf['descricao'])
        celula_desc.alignment = data_alignment
        celula_desc.border = border_thin

        # Coluna D: Regras
        celula_regras = ws.cell(row=row_num, column=4, value=rf['regras'])
        celula_regras.alignment = data_alignment
        celula_regras.border = border_thin

        # Coluna E: Processo (NOVA)
        celula_processo = ws.cell(row=row_num, column=5, value=rf['processo'])
        celula_processo.alignment = Alignment(horizontal="center", vertical="top")
        celula_processo.border = border_thin

        # Coluna F: Jornada (NOVA - renomeada de Epic/Grupo)
        celula_jornada = ws.cell(row=row_num, column=6, value=rf['jornada'])
        celula_jornada.alignment = data_alignment
        celula_jornada.border = border_thin

        # Colunas G-J: Editáveis pelo cliente (fundo amarelo)
        for col_num in range(7, 11):
            celula = ws.cell(row=row_num, column=col_num)
            celula.fill = editable_fill
            celula.alignment = data_alignment
            celula.border = border_thin

    # Ajustar larguras das colunas
    ws.column_dimensions['A'].width = 12   # Cód.
    ws.column_dimensions['B'].width = 35   # Nome Funcionalidade
    ws.column_dimensions['C'].width = 80   # Descrição (aumentado para texto completo)
    ws.column_dimensions['D'].width = 70   # Regras
    ws.column_dimensions['E'].width = 18   # Processo (novo)
    ws.column_dimensions['F'].width = 30   # Jornada (renomeado)
    ws.column_dimensions['G'].width = 15   # Prioridade
    ws.column_dimensions['H'].width = 18   # Status Discussão
    ws.column_dimensions['I'].width = 20   # Continua no sistema?
    ws.column_dimensions['J'].width = 40   # Notas

    # Congelar painéis (linha 1 fixa)
    ws.freeze_panes = 'A2'

    # Ajustar altura da linha de cabeçalho
    ws.row_dimensions[1].height = 30

    # Salvar planilha
    wb.save(output_path)
    print(f"[OK] Planilha gerada com sucesso: {output_path}")
    print(f"  - Total de linhas: {len(rfs) + 1} (1 cabeçalho + {len(rfs)} RFs)")
    print(f"  - Colunas: {len(cabecalhos)}")


def main():
    """Função principal do script."""
    documentacao_path = r"D:\IC2_Governanca\documentacao"
    output_path = r"D:\IC2_Governanca\funcionalidades.xlsx"

    print("\n" + "=" * 80)
    print("GERADOR DE PLANILHA DE FUNCIONALIDADES - ICONTROLIT")
    print("Versão 4.3 - Textos Completos (sem truncamento)")
    print("=" * 80 + "\n")

    # Extrair RFs
    rfs = extrair_rfs_yaml(documentacao_path)

    if not rfs:
        print("[ERRO] Nenhum RF encontrado.")
        return 1

    # Exibir amostra
    print("\n[INFO] Amostra dos primeiros 3 RFs:")
    for rf in rfs[:3]:
        # Encode/decode para ASCII para evitar erros de encoding no print
        nome_safe = rf['nome'][:60].encode('ascii', 'ignore').decode('ascii') if rf['nome'] else '(sem nome)'
        desc_safe = rf['descricao'][:60].encode('ascii', 'ignore').decode('ascii') if rf['descricao'] else '(sem descricao)'
        processo_safe = rf['processo'] if rf['processo'] else '(sem processo)'
        jornada_safe = rf['jornada']

        print(f"\n  RF: {rf['rf_id']}")
        print(f"  Nome: {nome_safe}...")
        print(f"  Descricao: {desc_safe}...")
        print(f"  Processo: {processo_safe}")
        print(f"  Jornada: {jornada_safe}")

    print("\n" + "-" * 80)

    # Gerar planilha
    gerar_planilha_excel(rfs, output_path)

    print("\n" + "=" * 80)
    print("CONCLUIDO COM SUCESSO")
    print("=" * 80 + "\n")

    return 0


if __name__ == '__main__':
    exit(main())
